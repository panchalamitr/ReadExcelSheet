import openpyxl
from pathlib import Path

xlsx_file = Path('PATH-TO FOLDER', 'FILE-NAME.xlsx')
wb_obj = openpyxl.load_workbook(xlsx_file)
 

# Read the active sheet:
sheet = wb_obj.active

# for row in sheet.iter_rows():
#     for column in range(sheet.max_column):
#         print(row[column].value)
# for column in range(sheet.max_column):
#     print(f'i = {column}')
#     for row in sheet.iter_rows():
#        if row[column].value is not None:
#             print(row[column].value)
org_string = 'Row: '
for row in sheet.iter_rows():
    for column in range(sheet.max_column):
       if row[column].value is not None:
          print(row[column].value)
          if column == 9 or column == 13 or column == 15 or column == 38:
            continue
          org_string += row[column].value + ", "

print(org_string)
org_string = ''
